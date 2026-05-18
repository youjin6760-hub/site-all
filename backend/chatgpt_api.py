import base64
import json
import mimetypes
import os
import re
import unicodedata

from pathlib import Path
import time
from typing import Any
from concurrent.futures import ThreadPoolExecutor, as_completed

from openai import OpenAI, RateLimitError
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from prompt_builder import (
    build_review_prompt,
    is_check_enabled,
    merge_review_checks,
    should_run_content_review,
    should_run_format_review,
)


ROOT = Path(os.getenv("APP_ROOT", Path(__file__).resolve().parent)).resolve()

RAW_DIR = Path(os.getenv("RAW_DIR", ROOT / "raw")).resolve()

REVIEWED_DIR = Path(os.getenv("REVIEWED_DIR", ROOT / "reviewed_json")).resolve()

# ISSUE_XLSX_PATH = ROOT / "json_issue_index.xlsx"
ISSUE_XLSX_PATH = Path(os.getenv("ISSUE_XLSX_PATH", ROOT / "chatgpt.xlsx")).resolve()

MODEL_NAME = os.getenv("CHATGPT_MODEL", "gpt-5.4")
MAX_TOKENS = int(os.getenv("CHATGPT_MAX_OUTPUT_TOKENS", "6000"))
CHATGPT_REASONING_EFFORT = os.getenv("CHATGPT_REASONING_EFFORT", "medium")

MAX_QUESTION_IMAGES = 2
MAX_CHOICE_IMAGES = 4
MAX_EXPLANATION_IMAGES = 2
MAX_PER_QUESTION_API_RETRIES = 3
MAX_CONSECUTIVE_REVIEW_FAILURES = 5

MAX_REVIEW_WORKERS = int(os.getenv("MAX_REVIEW_WORKERS", "3"))

def get_selected_review_checks(review_checks: dict[str, Any] | None = None) -> dict[str, dict[str, bool]]:
    """프론트에서 전달한 선택 검수값을 기본값과 병합합니다."""
    return merge_review_checks(review_checks)

def make_empty_review_result(question_id: str) -> dict:
    return {
        "question_id": question_id,
        "content_issues": [],
        "format_issues": [],
        "summary": {
            "has_issue": False,
            "issue_count": 0,
        },
    }


def guess_media_type(path: Path) -> str:
    mime, _ = mimetypes.guess_type(str(path))
    if mime in {"image/png", "image/jpeg", "image/gif", "image/webp"}:
        return mime
    return "image/png"


def make_text_block(text: str) -> dict:
    return {
        "type": "input_text",
        "text": text
    }


def make_image_block(path_str: str, label: str) -> list[dict]:
    path = Path(path_str)

    if not path.exists():
        return [make_text_block(f"[이미지 파일 누락] {label}: {path_str}")]

    media_type = guess_media_type(path)

    with open(path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")

    return [
        make_text_block(f"[첨부 이미지] {label}: {path.name}"),
        {
            "type": "input_image",
            "image_url": f"data:{media_type};base64,{encoded}"
        }
    ]
    
def safe_sheet_name(name: str) -> str:
    if not name:
        return "issues"

    name = str(name)

    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, "_")

    return name[:31]


def safe_excel_text(v):
    if v is None:
        return ""

    v = str(v)
    v = unicodedata.normalize("NFC", v)

    # 엑셀에서 문제 되는 제어문자만 제거
    v = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]", "", v)

    # 줄바꿈은 셀 깨짐 방지용으로 공백 처리
    v = v.replace("\r", " ").replace("\n", " ")

    return v.strip()


ISSUE_HEADERS = [
    "question_id",
    "file_name",
    "question_no",
    "subject_name",
    "sub_title",
    "issue_area",
    "issue_type",
    "reason",
    "suggestion",
]

def last_non_empty_row(ws):
    for row in range(ws.max_row, 0, -1):
        if any(
            ws.cell(row=row, column=col).value not in (None, "")
            for col in range(1, ws.max_column + 1)
        ):
            return row
    return 0


def save_issue_rows_to_xlsx(issue_rows):
    if not issue_rows:
        print("[ISSUE XLSX] 저장할 오류 없음")
        return

    is_existing_file = ISSUE_XLSX_PATH.exists()

    if is_existing_file:
        wb = load_workbook(ISSUE_XLSX_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    ws_map = {ws.title: ws for ws in wb.worksheets}
    existing_sheet_names = set(ws_map.keys())
    separated_sheets = set()

    for row_data in issue_rows:
        sheet_base = row_data.get("sub_title") or row_data.get("subject_name") or "issues"
        sheet_name = safe_sheet_name(sheet_base)

        if sheet_name not in ws_map:
            ws = wb.create_sheet(sheet_name)
            ws_map[sheet_name] = ws

            for col_idx, header in enumerate(ISSUE_HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            next_row = 2
        else:
            ws = ws_map[sheet_name]
            last_row = last_non_empty_row(ws)

            if (
                sheet_name in existing_sheet_names
                and last_row > 1
                and sheet_name not in separated_sheets
            ):
                next_row = last_row + 2
                separated_sheets.add(sheet_name)
            else:
                next_row = last_row + 1

        row_values = [
            row_data["question_id"],
            row_data["file_name"],
            row_data["question_no"],
            row_data["subject_name"],
            row_data["sub_title"],
            row_data["issue_area"],
            row_data["issue_type"],
            row_data["reason"],
            row_data["suggestion"],
        ]

        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=next_row, column=col_idx, value=safe_excel_text(value))

    wb.save(ISSUE_XLSX_PATH)
    print(f"[ISSUE XLSX 저장 완료] {ISSUE_XLSX_PATH}")
    

def is_cancel_exception(error: Exception) -> bool:
    return error.__class__.__name__ == "JobCancelled"

def is_fatal_review_error(error: Exception) -> bool:
    """
    스킵하면 안 되는 시스템성 오류를 구분합니다.

    이런 오류는 개별 문제 문제가 아니라 설정/인증/모델 문제일 가능성이 크므로
    즉시 중단해서 app.py에서 partial_failed 또는 failed 처리되게 합니다.
    """
    error_name = error.__class__.__name__
    text = str(error).lower()

    fatal_error_names = {
        "AuthenticationError",
        "PermissionDeniedError",
        "NotFoundError",
    }

    if error_name in fatal_error_names:
        return True

    fatal_keywords = [
        "api key",
        "invalid api key",
        "incorrect api key",
        "authentication",
        "permission",
        "insufficient_quota",
        "billing",
        "model",
        "does not exist",
        "not found",
        "unsupported model",
    ]

    return any(keyword in text for keyword in fatal_keywords)


def save_review_errors(skipped_errors: list[dict[str, Any]]) -> None:
    if not skipped_errors:
        return

    error_path = REVIEWED_DIR.parent / "review_errors.json"

    with open(error_path, "w", encoding="utf-8") as f:
        json.dump(skipped_errors, f, ensure_ascii=False, indent=2)

    print(f"[REVIEW 오류 목록 저장] {error_path}")
           
# =========================
# API 입력 구성
# =========================
def build_user_content(prompt: str, question_data: dict, mode: str, review_checks: dict[str, Any] | None = None) -> list[dict]:
    blocks: list[dict] = []

    data = question_data.get("data", {}) or {}
    image_elements = data.get("image_elements", []) or []
    explanation_images = data.get("explanation_images", []) or []

    has_question_image = bool(data.get("has_question_image"))
    has_choice_image = bool(data.get("has_choice_image"))
    has_image = bool(data.get("has_image"))
    body = data.get("body", "")
    extra_text = data.get("extra_text", "")
    choices = data.get("choices", [])
    selected_checks = get_selected_review_checks(review_checks)
    run_content = should_run_content_review(selected_checks)
    run_format = should_run_format_review(selected_checks)

    question_material_images = [
        item for item in image_elements
        if item.get("location") == "question"
    ]

    choice_material_images = [
        item for item in image_elements
        if item.get("location") == "choice"
    ]

    review_json = json.dumps({
        "problem": {
            "question_text": data.get("body", ""),
            "given_text": data.get("extra_text", ""),
            "given_images": question_material_images,
        },
        "choices": data.get("choices", []),
        "choice_images": choice_material_images,
        "answer": data.get("answer", ""),
        "keywords": data.get("keywords", data.get("keyword", "")),
        "explanation": data.get("explanation", ""),
        "explanation_images": data.get("explanation_images", []),
        "explanation_capture_meta": data.get("explanation_capture_meta", {}),
    }, ensure_ascii=False, indent=2)

    choice_image_lines = []
    for item in choice_material_images:
        label = item.get("caption_or_near_text", "") or ""
        file_name = Path(item.get("saved_path", "")).name
        original_or_near = item.get("ocr_or_extracted_text", "") or ""

        choice_image_lines.append(
            f"- {label or 'choice'}: 첨부 이미지 파일={file_name}, 원본/근처텍스트={original_or_near}"
        )

    choice_image_reference_text = (
        "\n".join(choice_image_lines)
        if choice_image_lines
        else "선지 이미지 없음"
    )
        
    matching_parts = []

    if (
        selected_checks["content"].get("choice_explanation_match")
        or selected_checks["format"].get("choice_explanation_exists")
        or selected_checks["format"].get("choice_explanation_format")
    ):
        matching_parts.append(
            "[선지/보기 해설 참고]\n"
            "- choices와 explanation은 위 [문제 데이터 JSON]에 이미 포함되어 있습니다.\n"
            "- 일반 문제는 choices와 explanation의 선지별 해설을 번호 기준으로 확인하세요.\n"
            "- 보기제시형(ㄱ/ㄴ/ㄷ/ㄹ, A/B/C 조합형)은 선택지 조합별 해설이 없어도 보기별 해설이 있으면 정상으로 보세요.\n"
        )
    else:
        matching_parts.append(
            "[선지 정보]\n"
            "- 필요한 경우 위 [문제 데이터 JSON]의 choices와 explanation을 참고하세요.\n"
        )

    if question_material_images:
        matching_parts.append(
            "[보기/제시자료 이미지 특별 검수 기준]\n"
            "- 이 문제는 보기/제시자료가 이미지로 제공될 수 있습니다.\n"
            "- problem.given_text가 비어 있거나 일부만 있어도, problem.given_images에 첨부된 이미지를 실제 보기/제시자료 원문으로 간주하세요.\n"
            "- 보기 이미지 안의 표, 행, 열, 값, 조건, 코드, 수식, ERD, 그림을 먼저 읽고 문제 성립 여부와 정답/해설을 검수하세요.\n"
            "- 해설이 보기 이미지의 값, 조건, 표 구조, 행/열 결과와 다르면 내용 오류 또는 해설 내용 오류로 기록하세요.\n"
            "- 이미지 내용을 확정적으로 읽을 수 없으면 임의로 내용 오류를 만들지 마세요. 단, 이미지가 없거나 잘못 캡처되어 문제 풀이 자체가 불가능하면 '문제 성립 오류'로 기록하세요.\n"
            "- 보기 이미지가 문제 풀이의 핵심인데 첨부되지 않았거나 잘못 캡처되었으면 문제 성립 오류로 기록하세요.\n"
        )

    if choice_material_images:
        matching_parts.append(
            "[이미지 선지 특별 검수 기준]\n"
            "- 이 문제는 선택지 자체가 이미지일 수 있습니다.\n"
            "- choices 배열의 '[이미지 선지: 파일명]'은 실제 선지 내용이 아니라 표시용 파일명입니다.\n"
            "- 이미지 선지가 있는 경우, choices 텍스트가 아니라 첨부된 선지 이미지를 실제 선지 원문으로 간주하세요.\n"
            "- 각 선지 이미지의 표/행/열/값을 먼저 읽고, explanation의 선지별 설명과 번호 기준으로 비교하세요.\n"
            "- SQL 실행 결과표, MERGE/UPDATE/INSERT/DELETE 결과표 이미지는 행 식별자와 갱신 대상 컬럼을 기준으로 행 단위 비교하세요.\n"
            "- 해설이 이미지에 없는 값, 누락된 행, 잘못된 결과, 다른 선지의 내용을 설명하면 '선지-해설 불일치'로 기록하세요.\n"
            "- 이미지에 정상적으로 존재하는 행을 해설이 오류라고 설명하거나, 이미지의 실제 오류 행/값과 다른 행/값을 오답 근거로 설명하면 '선지-해설 불일치'로 기록하세요.\n"
            "- 정답 번호가 맞더라도, 해설이 해당 이미지 선지의 오답 근거를 잘못 설명하면 '선지-해설 불일치'로 기록하세요.\n"
            "- 이미지에서 일부 행/값이 명확히 읽히면 그 읽힌 범위는 확정 근거로 사용하세요. 이미지 전체가 완벽히 읽히지 않는다는 이유로 명확히 보이는 행/값 비교를 생략하지 마세요.\n"
            "- 정답 검수도 이미지 선지의 실제 내용 기준으로 판단하세요.\n"
            "- 이미지 내용을 확정적으로 읽을 수 없으면 임의로 내용 오류를 만들지 마세요. 단, 이미지가 없거나 잘못 캡처되어 문제 풀이 자체가 불가능하면 '문제 성립 오류'로 기록하세요.\n"
            "[선지 이미지 매핑]\n"
            f"{choice_image_reference_text}\n"
        )

    matching_instruction = "\n".join(matching_parts)

    question_items = question_material_images
    choice_items = choice_material_images

    extra_instruction_parts = []

    if mode == "content" or (mode == "selected" and run_content):
        extra_instruction_parts.append(
            "[내용 검수 추가 지시]\n"
            "- 텍스트 문제라도 선택된 내용 검수 항목은 반드시 수행하세요.\n"
            "- 형식 오류보다 문제 자체 오류를 우선 검수하세요.\n"
            "- 보기/제시자료 이미지가 문제 풀이에 필수인데 없거나 잘못된 경우 반드시 문제 성립 오류로 판단하세요.\n"
            "- 보기/제시자료 이미지, 선지 이미지, 해설 스크린샷은 서로 독립적으로 검수하세요.\n"
            "- 보기/제시자료 이미지가 있는 경우 problem.given_text만 보지 말고 첨부된 problem.given_images를 실제 보기/자료 원문으로 판단하세요.\n"
            "- 보기 이미지의 표/행/열/값/조건과 해설 또는 정답 판단이 다르면 내용 오류로 기록하세요.\n"
            "- 이미지 선지가 있는 경우 choices의 파일명 텍스트가 아니라 첨부된 선지 이미지를 실제 선지 내용으로 판단하세요.\n"
            "- SQL 결과표 이미지 선지는 EMPNO, ID, KEY 같은 행 식별자와 SAL, AMOUNT, SCORE 같은 결과값 컬럼을 기준으로 해설의 오답 근거와 비교하세요.\n"
            "- 이미지 내용을 확정적으로 읽을 수 없으면 임의로 내용 오류를 만들지 마세요. 단, 이미지가 없거나 잘못 캡처되어 문제 풀이 자체가 불가능하면 '문제 성립 오류'로 기록하세요.\n"
            "- 키워드 검수가 선택된 경우 keywords가 문제의 핵심 개념과 맞는지, 누락되었거나 지나치게 넓지 않은지 확인하세요.\n"
            "- 문제 풀이가 불가능하면 다른 판단 없이 즉시 문제 성립 오류로 기록하세요.\n"
        )

    if mode == "format" or (mode == "selected" and run_format):
        extra_instruction_parts.append(
            "[형식 검수 추가 지시]\n"
            "- 내용 정오, 정답 정합성, 문제 성립 여부는 내용 검수 항목이 선택된 경우에만 판단하세요.\n"
            "- 선택된 형식 검수 항목만 판단하세요.\n"
            "- ㄱ/ㄴ/ㄷ/ㄹ 보기제시형 문제는 선지별 조합 해설이 없어도 보기별 해설이 있으면 정상으로 보세요.\n"
            "- 이미지 선지의 표/값/행 내용은 형식 검수에서 판단하지 마세요.\n"
            "- 이미지 선지의 파일명, 백틱, 캡션, OCR성 텍스트를 근거로 표현 오류를 만들지 마세요.\n"
            "- 이미지 선지가 있는 문제는 해설의 번호 구조와 문장 형식만 확인하고, 이미지 내용의 정오는 내용 검수로만 판단하세요.\n"
            "- 형식 검수에서는 이미지의 내용 정오를 판단하지 말고, 필요한 경우 해설 구조 확인 참고용으로만 사용하세요.\n"
            "- 단, 내용 검수 항목이 함께 선택된 경우 이미지의 실제 내용 판단은 위 [내용 검수 추가 지시]를 우선 적용하세요.\n"
        )

    if not extra_instruction_parts:
        extra_instruction_parts.append(
            "[검수 항목 없음]\n"
            "- 선택된 검수 항목이 없으므로 content_issues와 format_issues를 빈 배열로 반환하세요.\n"
        )

    extra_instruction = "\n".join(extra_instruction_parts)
        
    blocks.append(
        make_text_block(
            f"{prompt}\n\n"
            f"[문제 데이터 JSON]\n"
            f"{review_json}\n\n"
            f"[자료 구분]\n"
            f"- problem.question_text: 문제 질문입니다.\n"
            f"- problem.given_text: 보기/제시문 텍스트입니다.\n"
            f"- problem.given_images: 보기/제시자료 이미지입니다. 표, 테이블, ERD, 그림, 수식 이미지가 여기에 포함될 수 있습니다.\n"
            f"- choices: 정답 선택지입니다.\n"
            f"- choice_images: 선지 자체가 이미지인 경우의 선지 이미지입니다.\n"
            f"- explanation: 해설입니다.\n"
            f"- keywords: 문제 분류용 키워드입니다. 키워드 검수가 선택된 경우 문제 핵심 개념과 일치하는지 확인하세요.\n\n"
            f"{matching_instruction}"
            f"[검수용 추가 정보]\n"
            f"{extra_instruction}"
            f"- has_image: {has_image}\n"
            f"- has_question_image: {has_question_image}\n"
            f"- has_choice_image: {has_choice_image}\n"
            f"- question_image_count: {len(question_items)}\n"
            f"- choice_image_count: {len(choice_items)}\n"
            f"- explanation_image_count: {len(explanation_images)}\n"
            f"- body: {body}\n"
            f"- extra_text: {extra_text}\n"
            f"- choices_count: {len(choices)}\n"
            f"- 설명 문장, 마크다운, 코드블록 없이 순수 JSON만 반환하세요.\n"
        )
    )

    content_checks = selected_checks.get("content", {})
    format_checks = selected_checks.get("format", {})

    needs_given_image_review = (
        mode == "content"
        or review_checks is None
        or content_checks.get("problem_validity")
        or content_checks.get("image_validation")
        or content_checks.get("answer_validation")
        or content_checks.get("answer_correctness")
        or content_checks.get("explanation_logic")
    )

    needs_choice_image_review = (
        mode == "content"
        or review_checks is None
        or content_checks.get("problem_validity")
        or content_checks.get("image_validation")
        or content_checks.get("answer_validation")
        or content_checks.get("answer_correctness")
        or content_checks.get("choice_explanation_match")
    )

    needs_explanation_image_review = (
        mode in {"content", "format"}
        or review_checks is None
        or content_checks.get("explanation_logic")
        or content_checks.get("choice_explanation_match")
        or content_checks.get("expression_error")
        or content_checks.get("image_validation")
        or format_checks.get("long_explanation_manual_check")
    )

    include_question_images = bool(question_items and needs_given_image_review)
    include_choice_images = bool(choice_items and needs_choice_image_review)
    include_explanation_images = bool(explanation_images and needs_explanation_image_review)

    if include_question_images:
        if question_items:
            blocks.append(make_text_block("[보기/제시자료 이미지 검수 대상]"))
            for idx, item in enumerate(question_items[:MAX_QUESTION_IMAGES], start=1):
                saved_path = item.get("saved_path", "")
                near = item.get("caption_or_near_text", "") or item.get("ocr_or_extracted_text", "")
                label = f"보기/제시자료 이미지 {idx}"
                if near:
                    label += f" / {near}"
                blocks.extend(make_image_block(saved_path, label))
        else:
            blocks.append(make_text_block("[보기/제시자료 이미지 없음]"))
    else:
        blocks.append(make_text_block("[보기/제시자료 이미지 첨부 생략] 이미지 관련 내용 검수가 선택되지 않았습니다."))

    if include_choice_images:
        if choice_items:
            blocks.append(make_text_block("[선지 이미지 검수 대상]"))
            for idx, item in enumerate(choice_items[:MAX_CHOICE_IMAGES], start=1):
                saved_path = item.get("saved_path", "")
                near = item.get("caption_or_near_text", "") or item.get("ocr_or_extracted_text", "")
                label = f"선지 이미지 {idx}"
                if near:
                    label += f" / {near}"
                blocks.extend(make_image_block(saved_path, label))
        else:
            blocks.append(make_text_block("[선지 이미지 없음]"))
    else:
        blocks.append(make_text_block("[선지 이미지 첨부 생략] 선지 이미지 관련 검수가 선택되지 않았습니다."))

    if include_explanation_images:
        if explanation_images:
            blocks.append(make_text_block("[해설 스크린샷 검수 대상]"))
            for idx, path_str in enumerate(explanation_images[:MAX_EXPLANATION_IMAGES], start=1):
                blocks.extend(make_image_block(str(path_str), f"해설 스크린샷 {idx}"))
        else:
            blocks.append(make_text_block("[해설 스크린샷 없음]"))
    else:
        blocks.append(make_text_block("[해설 스크린샷 첨부 생략] 해설 이미지/긴 해설 검수가 선택되지 않았습니다."))

    return blocks


def extract_json_from_text(text: str) -> dict:
    text = text.strip()

    # 코드블록 제거
    text = re.sub(r"^```(?:json|JSON)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)

    # 1차: 전체가 JSON이면 바로 파싱
    try:
        return json.loads(text)
    except Exception:
        pass

    # 2차: 앞뒤에 설명이 붙어도 첫 JSON 객체만 찾아서 파싱
    decoder = json.JSONDecoder()

    for idx, ch in enumerate(text):
        if ch == "{":
            try:
                obj, end = decoder.raw_decode(text[idx:])
                return obj
            except Exception:
                continue

    raise ValueError("응답에서 JSON 객체를 찾지 못했습니다.")


def parse_review_response(text: str, question_id: str, mode: str, fallback_issue_area: str | None = None) -> dict:
    try:
        return extract_json_from_text(text)

    except Exception as e:
        DEBUG_DIR = REVIEWED_DIR.parent / "debug_api_response"
        DEBUG_DIR.mkdir(parents=True, exist_ok=True)

        debug_path = DEBUG_DIR / f"{question_id}_{mode}_parse_fail.txt"

        with open(debug_path, "w", encoding="utf-8") as f:
            f.write(text)

        issue_area = fallback_issue_area or ("content" if mode == "content" else "format")
        issue = {
            "type": "기타 내용 오류" if issue_area == "content" else "기타 형식 오류",
            "reason": f"검수 결과를 JSON으로 파싱하지 못했습니다. 원문 응답 저장: {debug_path.name}",
            "suggestion": "debug_api_response 폴더의 원문 응답을 확인하세요.",
        }

        return {
            "question_id": question_id,
            "content_issues": [issue] if issue_area == "content" else [],
            "format_issues": [issue] if issue_area == "format" else [],
            "summary": {
                "has_issue": True,
                "issue_count": 1,
            }
        }


VALID_REVIEW_ISSUE_TYPES = {
    "문제 성립 오류",
    "정답 불일치",
    "해설 내용 오류",
    "선지-해설 불일치",
    "표현/렌더링 오류",
    "키워드 오류",
    "기타 내용 오류",
    "해설 시작 형식 오류",
    "선지별 해설 누락",
    "선지 해설 형식 오류",
    "정답 문장 중복",
    "결론 누락",
    "최종 문장 형식 오류",
    "기타 형식 오류",
    "긴 해설 수동 검토 필요",
}


def is_false_positive_normal_issue(issue: dict) -> bool:
    issue_type = str(issue.get("type", "")).strip()
    reason = str(issue.get("reason", ""))
    suggestion = str(issue.get("suggestion", ""))
    text = f"{issue_type} {reason} {suggestion}"

    # 실제 허용 오류유형이면 정상 판단 문구가 일부 섞여 있어도 제거하지 않습니다.
    # 예: "정답은 올바르지만 해설이 틀림" 같은 내용 오류를 보존합니다.
    if issue_type in VALID_REVIEW_ISSUE_TYPES:
        normal_type_phrases = [
            "오류 없음",
            "정상",
            "해당 없음",
            "문제 없음",
            "문제 성립 오류 없음",
        ]
        return any(p in issue_type for p in normal_type_phrases)

    normal_phrases = [
        "오류는 없습니다",
        "형식 오류는 없습니다",
        "내용 오류는 없습니다",
        "문제는 없습니다",
        "이상 없습니다",
        "정상입니다",
        "일치합니다",
        "불일치는 없습니다",
        "불일치가 없습니다",
        "해당 없음",
        "수정할 필요 없습니다",
        "유지해도 됩니다",
        "문제 없어 보입니다",
        "확인 가능한 오류는 없습니다",
        "오류로 판단하지 않습니다",
        "오류로 보지 않습니다",
    ]

    return any(p in text for p in normal_phrases)


def normalize_issue_key(issue: dict) -> str:
    issue_type = str(issue.get("type", ""))
    reason = str(issue.get("reason", ""))
    suggestion = str(issue.get("suggestion", ""))

    text = f"{issue_type} {reason} {suggestion}"

    # 정답 문장 중복 묶기 - 해설 시작 형식 오류보다 먼저 처리
    if (
        issue_type == "정답 문장 중복"
        or (
            "정답은" in text
            and any(w in text for w in ["반복", "중복", "다시 등장", "한 번 더"])
        )
    ):
        return "정답 문장 중복"

    # 해설 시작 형식 오류 묶기
    if (
        issue_type == "해설 시작 형식 오류"
        or "해설 시작" in text
        or "explanation 시작" in text
        or "첫 문장" in text
        or "답은" in text
        or "정답은 X번입니다" in text
        or "정답은 [숫자]번입니다" in text
    ):
        return "해설 시작 형식 오류"

    # 선지 해설 형식 오류 묶기
    if (
        issue_type == "선지 해설 형식 오류"
        or "선지 해설" in text
        or "X. 선지 전체 문장" in text
    ):
        return "선지 해설 형식 오류"

    # 문체 오류 묶기
    if any(w in text for w in ["문체", "존댓말", "반말", "구어체", "평서체"]):
        return "문체 오류"

    # 정상 판단으로 작성된 issue 제거용 key
    if issue_type not in VALID_REVIEW_ISSUE_TYPES and any(w in text for w in [
        "오류는 없습니다",
        "형식 오류는 없습니다",
        "내용 오류는 없습니다",
        "문제는 없습니다",
        "이상 없습니다",
        "정상입니다",
        "일치합니다",
        "불일치는 없습니다",
        "불일치가 없습니다",
        "해당 없음",
        "해당 없음으로 보입니다",
        "수정할 필요 없습니다",
        "수정할 필요는 없습니다",
        "유지해도 됩니다",
        "문제 없어 보입니다",
        "문제는 없어 보입니다",
        "형식상 문제는 없어 보입니다",
        "확인 가능한 형식 오류는 없습니다",
        "확인 가능한 오류는 없습니다",
    ]):
        return "정상 판단"
    return f"{issue_type}|{reason[:50]}"


def dedupe_issues(issues: list) -> list:
    if not issues:
        return []

    result = []
    seen = set()

    for issue in issues:
        key = normalize_issue_key(issue)

        # 정상 판단은 오류로 저장하지 않음
        if key == "정상 판단":
            continue

        if key in seen:
            continue

        seen.add(key)
        result.append(issue)

    return result


def merge_selected_review_result(
    question_data: dict,
    selected_result: dict,
    review_checks: dict[str, Any] | None = None,
) -> dict:
    """선택형 통합 프롬프트 1회 호출 결과를 기존 reviewed JSON 구조로 맞춥니다."""
    checks = get_selected_review_checks(review_checks)
    run_content = any(checks["content"].values())
    run_format = any(checks["format"].values())

    content_issues = (selected_result.get("content_issues", []) or []) if run_content else []
    format_issues = (selected_result.get("format_issues", []) or []) if run_format else []

    content_issues = [issue for issue in content_issues if not is_false_positive_normal_issue(issue)]
    format_issues = [issue for issue in format_issues if not is_false_positive_normal_issue(issue)]

    content_issues = dedupe_issues(content_issues)
    format_issues = dedupe_issues(format_issues)

    has_issue = bool(content_issues or format_issues)

    return {
        **question_data,
        "content_issues": content_issues,
        "format_issues": format_issues,
        "summary": {
            "has_issue": has_issue,
            "issue_count": len(content_issues) + len(format_issues),
        },
    }


def refresh_review_summary(reviewed: dict) -> dict:
    content_issues = reviewed.get("content_issues", []) or []
    format_issues = reviewed.get("format_issues", []) or []

    reviewed.setdefault("summary", {})
    reviewed["summary"]["has_issue"] = bool(content_issues or format_issues)
    reviewed["summary"]["issue_count"] = len(content_issues) + len(format_issues)

    return reviewed


def filter_no_issue_entries(reviewed):
    """
    검수 항목에서 '오류 없음/정상 판단' 항목을 제거합니다.
    """
    def is_no_issue(issue):
        issue_type = str(issue.get("type", "")).strip()
        reason = str(issue.get("reason", ""))
        suggestion = str(issue.get("suggestion", ""))
        text = f"{issue_type} {reason} {suggestion}"

        if issue_type in VALID_REVIEW_ISSUE_TYPES:
            normal_type_phrases = [
                "오류 없음",
                "정상",
                "해당 없음",
                "문제 없음",
                "문제 성립 오류 없음",
            ]
            return any(p in issue_type for p in normal_type_phrases)

        no_issue_phrases = [
            "오류 없음",
            "오류는 없습니다",
            "형식 오류는 없습니다",
            "내용 오류는 없습니다",
            "문제는 없습니다",
            "이상 없습니다",
            "정상입니다",
            "불일치는 없습니다",
            "불일치가 없습니다",
            "수정할 필요 없습니다",
            "오류로 판단하지 않습니다",
            "오류로 보지 않습니다",
        ]

        return any(p in text for p in no_issue_phrases)

    reviewed["content_issues"] = [
        issue for issue in reviewed.get("content_issues", [])
        if not is_no_issue(issue)
    ]

    reviewed["format_issues"] = [
        issue for issue in reviewed.get("format_issues", [])
        if not is_no_issue(issue)
    ]

    return refresh_review_summary(reviewed)


def filter_quote_false_positive_issues(reviewed):
    data = reviewed.get("data", {}) or {}
    explanation = data.get("explanation", "") or ""
    choices = data.get("choices", []) or []
    answer = str(data.get("answer", "")).strip()

    correct_choice = ""
    if answer.isdigit():
        idx = int(answer) - 1
        if 0 <= idx < len(choices):
            correct_choice = str(choices[idx]).strip()

    last_line = explanation.strip().splitlines()[-1] if explanation.strip() else ""

    def is_quote_false_positive(issue):
        issue_type = str(issue.get("type", ""))
        text = (
            str(issue.get("reason", "")) + " " +
            str(issue.get("suggestion", ""))
        )

        quote_words = ["따옴표", "큰따옴표", "작은따옴표", "이스케이프", "충돌"]
        if not any(w in text for w in quote_words):
            return False

        # 내용 오류는 제거하지 않음. 형식 오류 중 결론 문장 따옴표 오판만 제거
        if issue_type not in ["최종 문장 형식 오류", "기타 형식 오류"]:
            return False

        # 결론 문장이 있고, 정답 선지 원문이 그대로 들어 있으면 따옴표 오판으로 봄
        if "따라서" in last_line and correct_choice and correct_choice in last_line:
            return True

        return False

    reviewed["format_issues"] = [
        issue for issue in reviewed.get("format_issues", [])
        if not is_quote_false_positive(issue)
    ]

    return refresh_review_summary(reviewed)

    
def add_duplicate_answer_sentence_issue(reviewed: dict, question_data: dict) -> dict:
    explanation = question_data.get("data", {}).get("explanation", "") or ""

    pattern = r"정답은\s*[1-5]\s*번입니다\.?"
    matches = list(re.finditer(pattern, explanation))

    if not matches:
        return reviewed

    # 공백 제거 후 첫 문장이 정답 문장으로 시작하는지 확인
    stripped = explanation.lstrip()
    first_match = matches[0]

    starts_correctly = first_match.start() == len(explanation) - len(stripped)

    # 오류 조건:
    # 1) 정답 문장이 2개 이상 있음
    # 2) 정답 문장이 1개만 있어도 첫 문장 위치가 아님
    if len(matches) == 1 and starts_correctly:
        return reviewed

    issue = {
        "type": "정답 문장 중복",
        "reason": "해설 첫 문장이 아닌 위치에 '정답은 X번입니다.' 형식의 문장이 등장하여 해설 형식 기준에 맞지 않습니다.",
        "suggestion": "해설 첫 문장을 정확히 '정답은 X번입니다.' 형식으로 작성하고, 해설 중간 또는 결론 문단 직전의 정답 문장은 삭제하세요.",
    }

    reviewed.setdefault("format_issues", [])

    already_exists = any(
        normalize_issue_key(i) == "정답 문장 중복"
        for i in reviewed.get("format_issues", [])
    )

    if not already_exists:
        reviewed["format_issues"].append(issue)

    # 혹시 후처리 후에도 중복이 생겼을 경우 한 번 더 정리
    reviewed["format_issues"] = dedupe_issues(reviewed.get("format_issues", []))

    return refresh_review_summary(reviewed)


def add_explanation_manual_review_issue(
    reviewed: dict,
    question_data: dict,
    selected_checks: dict[str, Any],
) -> dict:
    format_checks = selected_checks.get("format", {}) or {}

    if not format_checks.get("long_explanation_manual_check"):
        return reviewed

    data = question_data.get("data", {}) or {}
    meta = data.get("explanation_capture_meta", {}) or {}

    capture_mode = str(meta.get("capture_mode") or "")
    needs_manual_review = bool(meta.get("needs_manual_review"))

    manual_modes = {
        "too_long_after_zoom",
        "skipped",
        "failed",
    }

    if capture_mode not in manual_modes and not needs_manual_review:
        return reviewed

    # not_needed / not_attempted는 수동 검토 대상으로 보지 않습니다.
    if capture_mode in {"not_needed", "not_attempted"}:
        return reviewed

    capture_mode_reason_map = {
        "too_long_after_zoom": "해설이 화면에 한 번에 들어오지 않아 자동 캡처만으로 전체 해설을 확인하기 어렵습니다.",
        "skipped": "해설 캡처 전 화면 상태를 안정적으로 맞추지 못해 자동 캡처를 건너뛰었습니다.",
        "failed": "해설 영역을 자동으로 캡처하지 못했습니다.",
    }

    reason_text = capture_mode_reason_map.get(
        capture_mode,
        "해설 자동 캡처 결과 수동 확인이 필요한 상태로 기록되었습니다.",
    )

    issue = {
        "type": "긴 해설 수동 검토 필요",
        "reason": reason_text,
        "suggestion": (
            "관리자 화면에서 해설 전체가 누락 없이 표시되는지, "
            "수식·표·문자·렌더링이 깨지지 않았는지 직접 확인하세요."
        ),
    }

    reviewed.setdefault("format_issues", [])

    already_exists = any(
        str(i.get("type", "")) == "긴 해설 수동 검토 필요"
        for i in reviewed.get("format_issues", [])
    )

    if not already_exists:
        reviewed["format_issues"].append(issue)

    return refresh_review_summary(reviewed)


def review_one(client, prompt: str, question_data: dict, mode: str, review_checks: dict[str, Any] | None = None, fallback_issue_area: str | None = None) -> dict:
    content = build_user_content(prompt, question_data, mode, review_checks=review_checks)

    for attempt in range(5):
        try:
            resp = client.responses.create(
                model=MODEL_NAME,
                input=[
                    {
                        "role": "user",
                        "content": content,
                    }
                ],
                max_output_tokens=MAX_TOKENS,
                reasoning={
                    "effort": CHATGPT_REASONING_EFFORT
                }
            )
            break

        except RateLimitError:
            wait = 5 * (attempt + 1)
            print(f"[RateLimit] {wait}초 대기 후 재시도")
            time.sleep(wait)

        except Exception as e:
            print(f"[API 오류] {e}")
            raise

    else:
        raise RuntimeError("RateLimit 재시도 실패")

    text = getattr(resp, "output_text", "") or ""
    text = text.strip()

    return parse_review_response(text, question_data.get("question_id", ""), mode, fallback_issue_area=fallback_issue_area)

def review_one_with_retries(
    client,
    prompt: str,
    question_data: dict,
    mode: str,
    review_checks: dict[str, Any] | None = None,
    fallback_issue_area: str | None = None,
    max_retries: int = MAX_PER_QUESTION_API_RETRIES,
) -> dict:
    last_error: Exception | None = None

    for attempt in range(1, max_retries + 1):
        try:
            return review_one(
                client,
                prompt,
                question_data,
                mode=mode,
                review_checks=review_checks,
                fallback_issue_area=fallback_issue_area,
            )

        except Exception as e:
            if is_cancel_exception(e):
                raise

            if is_fatal_review_error(e):
                raise

            last_error = e

            if attempt < max_retries:
                wait = min(3 * attempt, 10)
                print(
                    f"[문제별 API 재시도] "
                    f"{question_data.get('question_id', '')} "
                    f"{attempt}/{max_retries} 실패: {e} / {wait}초 후 재시도"
                )
                time.sleep(wait)

    raise RuntimeError(f"문제별 API 재시도 {max_retries}회 실패: {last_error}") from last_error


def review_single_raw_file_task(
    raw_file: Path,
    api_key: str,
    selected_prompt: str,
    selected_checks: dict[str, Any],
    run_content: bool,
    run_format: bool,
    cancel_checker=None,
) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    """
    병렬 처리용 단일 문제 검수 함수입니다.
    여기서는 API 호출까지만 처리하고,
    reviewed_json 저장, issue_rows 정리, xlsx 저장은 메인 루프에서 처리합니다.
    """
    check_cancel(cancel_checker)

    with open(raw_file, "r", encoding="utf-8") as f:
        question_data = json.load(f)

    if not run_content and not run_format:
        selected_result = make_empty_review_result(
            question_data.get("question_id", "")
        )
    else:
        client = OpenAI(
            api_key=api_key,
            timeout=120,
        )

        fallback_area = "content" if run_content else "format"

        selected_result = review_one_with_retries(
            client,
            selected_prompt,
            question_data,
            mode="selected",
            review_checks=selected_checks,
            fallback_issue_area=fallback_area,
        )

    check_cancel(cancel_checker)

    return raw_file, question_data, selected_result


def question_no_from_filename(path: Path):
    try:
        return int(path.stem.split("_")[-1])
    except Exception:
        return 999999



def configure_review_dirs(
    raw_dir: str | Path | None = None,
    reviewed_dir: str | Path | None = None,
    issue_xlsx_path: str | Path | None = None,
):
    global RAW_DIR, REVIEWED_DIR, ISSUE_XLSX_PATH

    if raw_dir is not None:
        RAW_DIR = Path(raw_dir).resolve()
    if reviewed_dir is not None:
        REVIEWED_DIR = Path(reviewed_dir).resolve()
    if issue_xlsx_path is not None:
        ISSUE_XLSX_PATH = Path(issue_xlsx_path).resolve()

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    REVIEWED_DIR.mkdir(parents=True, exist_ok=True)
    ISSUE_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)


def review_raw_files(
    raw_files: list[str | Path] | None = None,
    reviewed_dir: str | Path | None = None,
    issue_xlsx_path: str | Path | None = None,
    write_excel: bool = True,
    cancel_checker=None,
    review_checks: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """
    전달받은 raw JSON 파일 목록만 ChatGPT로 검수합니다.

    실사용 정책:
    - 개별 문제 API 오류: 최대 3회 재시도 후 스킵
    - 스킵 문제: review_errors.json에 저장
    - 연속 5개 실패: 시스템성 문제 가능성이 높으므로 중단
    - 인증/API키/모델 설정 오류: 즉시 중단
    - 사용자 취소: 즉시 중단
    """
    configure_review_dirs(
        raw_dir=RAW_DIR,
        reviewed_dir=reviewed_dir,
        issue_xlsx_path=issue_xlsx_path,
    )

    load_dotenv(ROOT / ".env")

    api_key = os.getenv("CHATGPT_API_KEY")
    if not api_key:
        raise RuntimeError("환경변수 CHATGPT_API_KEY가 필요합니다.")

    selected_checks = get_selected_review_checks(review_checks)
    selected_prompt = build_review_prompt(selected_checks)
    run_content = should_run_content_review(selected_checks)
    run_format = should_run_format_review(selected_checks)

    issue_rows: list[dict[str, Any]] = []
    reviewed_results: list[dict[str, Any]] = []
    skipped_errors: list[dict[str, Any]] = []

    consecutive_failures = 0

    if raw_files is None:
        raw_paths = sorted(RAW_DIR.glob("*.json"), key=question_no_from_filename)
    else:
        raw_paths = sorted([Path(p) for p in raw_files], key=question_no_from_filename)

    if not raw_paths:
        print("[정보] 검수할 raw JSON이 없습니다.")
        return []

    review_payloads: dict[Path, tuple[dict[str, Any], dict[str, Any]]] = {}
    error_payloads: dict[Path, Exception] = {}

    try:
        worker_count = max(1, min(MAX_REVIEW_WORKERS, len(raw_paths)))

        if worker_count == 1:
            for raw_file in raw_paths:
                try:
                    _, question_data, selected_result = review_single_raw_file_task(
                        raw_file=raw_file,
                        api_key=api_key,
                        selected_prompt=selected_prompt,
                        selected_checks=selected_checks,
                        run_content=run_content,
                        run_format=run_format,
                        cancel_checker=cancel_checker,
                    )
                    review_payloads[raw_file] = (question_data, selected_result)

                except Exception as e:
                    error_payloads[raw_file] = e

        else:
            with ThreadPoolExecutor(max_workers=worker_count) as executor:
                future_map = {
                    executor.submit(
                        review_single_raw_file_task,
                        raw_file,
                        api_key,
                        selected_prompt,
                        selected_checks,
                        run_content,
                        run_format,
                        cancel_checker,
                    ): raw_file
                    for raw_file in raw_paths
                }

                for future in as_completed(future_map):
                    raw_file = future_map[future]

                    try:
                        _, question_data, selected_result = future.result()
                        review_payloads[raw_file] = (question_data, selected_result)

                    except Exception as e:
                        error_payloads[raw_file] = e

        for raw_file in raw_paths:
            check_cancel(cancel_checker)

            if raw_file in error_payloads:
                e = error_payloads[raw_file]

                if is_cancel_exception(e):
                    raise

                if is_fatal_review_error(e):
                    print(f"[치명적 API/설정 오류] {raw_file.name}: {e}")
                    raise e

                consecutive_failures += 1

                error_item = {
                    "file_name": raw_file.name,
                    "question_id": "",
                    "question_no": "",
                    "error_type": e.__class__.__name__,
                    "error_message": str(e),
                    "consecutive_failures": consecutive_failures,
                }
                skipped_errors.append(error_item)

                print(
                    f"[SKIP API 오류] {raw_file.name}: {e} "
                    f"/ 연속 실패 {consecutive_failures}개"
                )

                if consecutive_failures >= MAX_CONSECUTIVE_REVIEW_FAILURES:
                    raise RuntimeError(
                        f"API 검수 오류가 연속 {MAX_CONSECUTIVE_REVIEW_FAILURES}개 발생하여 "
                        f"작업을 중단합니다. 마지막 오류: {raw_file.name} / {e}"
                    ) from e

                continue

            question_data, selected_result = review_payloads[raw_file]

            try:
                check_cancel(cancel_checker)

                merged = merge_selected_review_result(
                    question_data,
                    selected_result,
                    selected_checks,
                )
                merged = filter_no_issue_entries(merged)
                merged = filter_quote_false_positive_issues(merged)

                if is_check_enabled(selected_checks, "format", "duplicate_answer_sentence"):
                    merged = add_duplicate_answer_sentence_issue(merged, question_data)

                merged = add_explanation_manual_review_issue(
                    merged,
                    question_data,
                    selected_checks,
                )

            except Exception as e:
                if is_cancel_exception(e):
                    raise

                if is_fatal_review_error(e):
                    print(f"[치명적 API/설정 오류] {raw_file.name}: {e}")
                    raise

                consecutive_failures += 1

                error_item = {
                    "file_name": raw_file.name,
                    "question_id": question_data.get("question_id", ""),
                    "question_no": question_data.get("question_no", ""),
                    "error_type": e.__class__.__name__,
                    "error_message": str(e),
                    "consecutive_failures": consecutive_failures,
                }
                skipped_errors.append(error_item)

                print(
                    f"[SKIP API 오류] {raw_file.name}: {e} "
                    f"/ 연속 실패 {consecutive_failures}개"
                )

                if consecutive_failures >= MAX_CONSECUTIVE_REVIEW_FAILURES:
                    raise RuntimeError(
                        f"API 검수 오류가 연속 {MAX_CONSECUTIVE_REVIEW_FAILURES}개 발생하여 "
                        f"작업을 중단합니다. 마지막 오류: {raw_file.name} / {e}"
                    ) from e

                continue

            # 성공하면 연속 실패 카운트 초기화
            consecutive_failures = 0

            out_path = REVIEWED_DIR / raw_file.name

            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(merged, f, ensure_ascii=False, indent=2)

            reviewed_results.append(merged)

            if merged["summary"].get("has_issue") is True:
                issues = []

                for issue in merged.get("content_issues", []):
                    issues.append(("content", issue))

                for issue in merged.get("format_issues", []):
                    issues.append(("format", issue))

                for issue_area, issue in issues:
                    row_data = {
                        "question_id": merged.get("question_id", ""),
                        "file_name": raw_file.name,
                        "question_no": merged.get("question_no", ""),
                        "subject_name": merged.get("subject_name", ""),
                        "sub_title": merged.get("sub_title", ""),
                        "issue_area": issue_area,
                        "issue_type": issue.get("type", ""),
                        "reason": issue.get("reason", ""),
                        "suggestion": issue.get("suggestion", ""),
                    }

                    print(
                        f"[ISSUE 행 추가] {raw_file.name} "
                        f"/ {issue_area} / {issue.get('type', '')}"
                    )
                    issue_rows.append(row_data)

                print(f"[ISSUE XLSX 기록] {raw_file.name}")

            print(
                f"[REVIEW 저장] {out_path.name} "
                f"/ issue={merged['summary'].get('has_issue')} "
                f"/ issue_count={merged['summary'].get('issue_count')}"
            )

    finally:
        save_review_errors(skipped_errors)

        if write_excel:
            save_issue_rows_to_xlsx(issue_rows)

    return reviewed_results


def check_cancel(cancel_checker=None):
    if cancel_checker is not None:
        cancel_checker()

def review_job_dir(
    job_dir: str | Path,
    write_excel: bool = True,
    cancel_checker=None,
    review_checks: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """job_dir/raw 안의 JSON만 검수하고 job_dir/reviewed_json에 결과를 저장합니다."""
    job_dir = Path(job_dir).resolve()
    raw_dir = job_dir / "raw"
    reviewed_dir = job_dir / "reviewed_json"

    configure_review_dirs(
        raw_dir=raw_dir,
        reviewed_dir=reviewed_dir,
        issue_xlsx_path=job_dir / "chatgpt.xlsx",
    )

    raw_files = sorted(raw_dir.glob("*.json"), key=question_no_from_filename)
    return review_raw_files(
        raw_files=raw_files,
        reviewed_dir=reviewed_dir,
        issue_xlsx_path=job_dir / "chatgpt.xlsx",
        write_excel=write_excel,
        cancel_checker=cancel_checker,
        review_checks=review_checks,
    )


def main():
    # 기존 로컬 실행 호환: RAW_DIR의 전체 JSON을 검수합니다.
    review_raw_files()


if __name__ == "__main__":
    main()
